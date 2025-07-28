# catxls.py
from openpyxl import load_workbook
import sys

if len(sys.argv) != 2:
    print("Usage: python catxls.py <file.xlsx>")
    sys.exit(1)

file = sys.argv[1]
wb = load_workbook(file, data_only=True)
ws = wb.active

for row in ws.iter_rows(values_only=True):
    print("\t".join(str(cell) if cell is not None else "" for cell in row))

