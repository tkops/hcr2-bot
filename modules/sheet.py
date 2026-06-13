#!/usr/bin/env python3
import sqlite3
from typing import Optional, List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Alignment
from pathlib import Path
from datetime import date
from hcr2.exporters import excel as excel_exporter
from hcr2.integrations.nextcloud import (
    NEXTCLOUD_BASE,
    delete_file,
    download_file,
    match_sheet_remote_path,
    upload_file,
)
from hcr2.services import sheets as sheet_service
from modules.common import (
    DB_PATH,
    connect_db,
    is_absent_on,
    is_help_request,
    parse_date_or_none,
    parse_int,
    print_command_help,
    print_unknown_command,
)

# --- Columns that are not exported/imported ---
EXCLUDED_PLAYER_COLS = {
    "created_at",
    "team",
    "away_from",
    "away_until",
    "active_modified",
    "about",
    "preferred_vehicles",
    "playstyle",
    "language",
    "country_code",
    "last_modified",
}

# --- Player export/import targets ---
PLAYERS_XLSX_NAME = "Ladys.xlsx"
PLAYERS_REMOTE_PATH = NEXTCLOUD_BASE / PLAYERS_XLSX_NAME
PLAYERS_LOCAL_TMP = Path("tmp") / PLAYERS_XLSX_NAME

# --- Donations export/import targets ---
DONATIONS_XLSX_NAME = "Donations.xlsx"
DONATIONS_REMOTE_PATH = NEXTCLOUD_BASE / DONATIONS_XLSX_NAME
DONATIONS_LOCAL_TMP = Path("tmp") / DONATIONS_XLSX_NAME


def sanitize_filename(s):
    return sheet_service.sanitize_filename(s)


def get_match_info(conn, match_id):
    c = conn.cursor()
    c.execute("""
        SELECT m.id, m.start, m.season_number, m.opponent, e.name
        FROM match m
        JOIN teamevent e ON m.teamevent_id = e.id
        WHERE m.id = ?
    """, (match_id,))
    return c.fetchone()


def get_active_players(conn) -> List[Tuple[int, str, Optional[str], Optional[str]]]:
    c = conn.cursor()
    c.execute("""
        SELECT id, name, away_from, away_until
        FROM players
        WHERE active = 1 AND team = 'PLTE'
        ORDER BY name
    """)
    return c.fetchall()


def _is_absent_on(match_day: date, frm: Optional[str], until: Optional[str]) -> bool:
    return is_absent_on(match_day, frm, until)


def upload_to_nextcloud(local_path, remote_path, *, overwrite: bool = False):
    return upload_file(local_path, remote_path, overwrite=overwrite)


def delete_from_nextcloud(remote_path) -> bool:
    return delete_file(remote_path)


def download_from_nextcloud(season, filename, local_path):
    return download_file(match_sheet_remote_path(season, filename), Path(local_path))


# -------------------- Ranking Logic --------------------

def _fetch_season_rows(conn: sqlite3.Connection, season_number: int):
    cur = conn.cursor()
    cur.execute("""
        SELECT
            ms.player_id,
            p.name,
            p.team,
            p.active,
            ms.score,
            m.id,
            t.tracks
        FROM matchscore ms
        JOIN players p ON ms.player_id = p.id
        JOIN match m ON ms.match_id = m.id
        JOIN teamevent t ON m.teamevent_id = t.id
        WHERE m.season_number = ?
    """, (season_number,))
    return cur.fetchall()


def rank_active_plte_for_season(conn: sqlite3.Connection, season_number: int) -> List[Tuple[int, str, Optional[str], Optional[str]]]:
    cur = conn.cursor()
    cur.execute("""
        SELECT id, name, away_from, away_until
        FROM players
        WHERE active = 1 AND team = 'PLTE'
    """)
    base_players = cur.fetchall()
    if not base_players:
        return []
    id_to_player = {pid: (pid, name, a_from, a_until) for (pid, name, a_from, a_until) in base_players}

    rows = _fetch_season_rows(conn, season_number)

    scores_by_match = {}
    for pid, name, team, active, score, match_id, tracks in rows:
        if team != "PLTE" or not active:
            continue
        if score is None:
            continue
        scaled = score * 4 / tracks if tracks else score
        scores_by_match.setdefault(match_id, []).append((pid, scaled))

    import statistics
    player_deltas, player_counts = {}, {}
    for _, entries in scores_by_match.items():
        vals = [s for _, s in entries]
        if not vals:
            continue
        try:
            med = statistics.median(vals)
        except statistics.StatisticsError:
            continue
        for pid, s in entries:
            delta = s - med
            player_deltas.setdefault(pid, []).append(delta)
            player_counts[pid] = player_counts.get(pid, 0) + 1

    with_scores, without_scores = [], []
    for pid, (pid_, name, a_from, a_until) in id_to_player.items():
        deltas = player_deltas.get(pid)
        if deltas:
            avg_delta = round(sum(deltas) / len(deltas))
            cnt = player_counts.get(pid, 0)
            with_scores.append((avg_delta, -cnt, name.lower(), (pid_, name, a_from, a_until)))
        else:
            without_scores.append((name.lower(), (pid_, name, a_from, a_until)))

    with_scores_sorted = [p for _, _, _, p in sorted(with_scores, key=lambda x: (x[0], x[1], x[2]), reverse=True)]
    without_scores_sorted = [p for _, p in sorted(without_scores, key=lambda x: x[0])]
    return with_scores_sorted + without_scores_sorted


# -------------------- Excel Generation & Import (Match Sheet) --------------------

def generate_excel(match, players, output_path):
    """
    Match sheet. Unchanged except for standard formatting.
    """
    match_id, match_date_str, season, opponent, event = match

    md = parse_date_or_none(match_date_str) or date(1970, 1, 1)

    filename = sheet_service.match_sheet_filename(match_id, event, opponent)
    filepath = sheet_service.match_sheet_local_path(output_path, season, filename)
    folder = filepath.parent

    folder.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Match Info"

    ws.append([f"Match ID: {match_id}", f"Date: {match_date_str}", f"Season: {season}", f"Opponent: {opponent}", f"Event: {event}"])

    ws.insert_rows(2, amount=1)
    ws["A2"] = "Result"
    ws["B2"] = "Power Ladies -->"
    ws["C2"] = ""
    ws["D2"] = ""
    ws["E2"] = f"<-- {opponent}"

    ws.append(["MatchID", "PlayerID", "Player", "Score", "Points", "Absent", "Checkin", "Notes"])

    for row in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=7):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 9
    ws.column_dimensions["H"].width = 130

    ws["H3"] = (
        "H1: Did not drive: enter Score=0 and Points=0.\n"
        "H2: Set Absent to true when a player is excused (vacation etc.).\n"
        "H3: Set Checkin to true when a player logged into the match but did not drive.\n"
        "H4: If a player left the team but is still listed, delete the row.\n"
        "H5: If a player is missing, add them with the correct ID.\n"
        "H6: If a missing player has not been created yet, enter 'a' for add in column B instead of the ID. The player is created during import.\n"
        "H7: Enter the match results in cell C2 (Ladies) and D2 (opponent)."
    )
    ws["H3"].alignment = Alignment(wrap_text=True, vertical="top")

    for pid, name, a_from, a_until in players:
        absent_flag = _is_absent_on(md, a_from, a_until)
        ws.append([match_id, pid, name, "", "", "true" if absent_flag else "false", "", ""])

    align_center = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = align_center

    wb.save(filepath)

    remote_path = sheet_service.match_sheet_remote_path_for_filename(season, filename)
    upload_to_nextcloud(filepath, remote_path)  # No overwrite for match sheets.

    try:
        filepath.unlink()
    except Exception:
        pass

    web_url = sheet_service.scores_web_url(season)
    return f"[{filename}]({web_url})", True


def import_excel_to_matchscore(match_id):
    with sqlite3.connect(DB_PATH) as conn:
        match = get_match_info(conn, match_id)
        if not match:
            print("❌ No match found.")
            return

        match_id, _, season, opponent, event = match
        filename = sheet_service.match_sheet_filename(match_id, event, opponent)
        local_path = Path("tmp") / filename

        local_path.parent.mkdir(parents=True, exist_ok=True)
        download_from_nextcloud(season, filename, local_path)

        ladyscore, oppscore, rows = excel_exporter.read_match_sheet_workbook(local_path)
        validation = sheet_service.validate_match_sheet_rows(
            lady_score=ladyscore,
            opponent_score=oppscore,
            rows=rows,
        )

        if validation.errors:
            print("❌ Import aborted due to validation errors:")
            for msg in validation.errors:
                print(" -", msg)
            return

        result = sheet_service.apply_match_sheet_entries(
            match_id=match_id,
            entries=validation.entries,
            score_ladys=ladyscore if ladyscore is not None else 0,
            score_opponent=oppscore if oppscore is not None else 0,
        )

        try:
            local_path.unlink()
        except Exception:
            pass

        web_url = sheet_service.scores_web_url(season)
        status = "Changed" if result.changed > 0 else "Unchanged"
        score_status = "Score updated" if result.score_updated else "Score update failed"
        print(
            f"✅ [{filename}]({web_url}) "
            f"({status}, {result.imported} imported, {result.changed} changed; {score_status})"
        )


# ===================== Players: Export/Import (active PLTE, excludes, formatting) =====================

def _download_players_xlsx(local_path: Path = PLAYERS_LOCAL_TMP) -> Optional[Path]:
    return download_file(PLAYERS_REMOTE_PATH, local_path)


def _upload_players_xlsx(local_path: Path):
    # Only the players workbook may be overwritten.
    return upload_to_nextcloud(local_path, PLAYERS_REMOTE_PATH, overwrite=True)


def export_players_to_excel(db_path: str = DB_PATH, out_path: Path = PLAYERS_LOCAL_TMP):
    export_data = sheet_service.get_player_export_data(db_path, excluded_columns=EXCLUDED_PLAYER_COLS)
    if export_data is None:
        print("❌ players table not found")
        return

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = excel_exporter.build_players_workbook(export_data.columns, export_data.rows)
    wb.save(out_path)

    url, created = _upload_players_xlsx(out_path)
    try:
        out_path.unlink()
    except Exception:
        pass

    web_url = sheet_service.scores_web_url()
    print(f"✅ [Power-Ladys-Scores/{PLAYERS_XLSX_NAME}]({web_url}) ({'Created' if created else 'Updated'})")


def import_players_from_excel(db_path: str = DB_PATH, local_xlsx: Optional[Path] = None):
    local = local_xlsx or _download_players_xlsx()
    if not local or not local.exists():
        print("❌ players Excel not found on Nextcloud")
        return

    header, workbook_rows = excel_exporter.read_players_workbook(local)
    if header is None or workbook_rows is None:
        print("❌ First row must contain column names including 'id'")
        return

    result = sheet_service.import_player_rows(
        db_path,
        workbook_rows,
        excluded_columns=EXCLUDED_PLAYER_COLS,
    )

    # Delete local copy on a best-effort basis.
    try:
        local.unlink()
    except Exception:
        pass

    # Delete only the players Excel file in Nextcloud; leave match sheets untouched.
    deleted = delete_from_nextcloud(PLAYERS_REMOTE_PATH)
    status = "deleted" if deleted else "delete failed"

    print(
        f"✅ players import: {result.updated} updated, {result.inserted} inserted, "
        f"{result.skipped} skipped, {result.errors} errors ({status} in Nextcloud)"
    )


# ===================== Donations: Export/Import =====================

def _download_donations_xlsx(local_path: Path = DONATIONS_LOCAL_TMP) -> Optional[Path]:
    return download_file(DONATIONS_REMOTE_PATH, local_path)


def _upload_donations_xlsx(local_path: Path):
    # Donations.xlsx may be overwritten.
    return upload_to_nextcloud(local_path, DONATIONS_REMOTE_PATH, overwrite=True)


def export_donations_to_excel(db_path: str = DB_PATH, out_path: Path = DONATIONS_LOCAL_TMP):
    rows = sheet_service.get_donation_export_rows(db_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    today_str = date.today().isoformat()
    wb = excel_exporter.build_donations_workbook(rows, today_str)
    wb.save(out_path)

    url, created = _upload_donations_xlsx(out_path)
    try:
        out_path.unlink()
    except Exception:
        pass

    web_url = sheet_service.scores_web_url()
    print(f"✅ [Power-Ladys-Scores/{DONATIONS_XLSX_NAME}]({web_url}) ({'Created' if created else 'Updated'})")

def import_donations_from_excel(db_path: str = DB_PATH, local_xlsx: Optional[Path] = None):
    local = local_xlsx or _download_donations_xlsx()
    if not local or not local.exists():
        print("❌ donations Excel not found on Nextcloud")
        return

    date_str, donation_entries, errors = excel_exporter.read_donations_workbook(local)
    if date_str is None:
        print("❌ No valid date in cell A2")
        return

    result = sheet_service.import_donation_entries(
        db_path,
        date_str,
        donation_entries,
        initial_errors=errors,
    )

    # Delete local copy.
    try:
        local.unlink()
    except Exception:
        pass

    # Delete the donations Excel file in Nextcloud, same as players.
    deleted = delete_from_nextcloud(DONATIONS_REMOTE_PATH)
    status = "deleted" if deleted else "delete failed"

    print(f"✅ donations import: {result.added} added, {result.errors} errors ({status} in Nextcloud)")


# ===================== CLI =====================

USAGE_SHEET_CREATE = "Usage: sheet create <match_id>"
USAGE_SHEET_IMPORT = "Usage: sheet import <match_id>"
USAGE_SHEET_PLAYER = "Usage: sheet player <export|import>"
USAGE_SHEET_DONATIONS = "Usage: sheet donations <export|import>"

def print_help():
    print_command_help(
        usage="hcr2.py sheet <command> [options]",
        commands=[
            ("create <match_id>", "Create one Excel file and upload it to Nextcloud"),
            ("import <match_id>", "Import scores from one Excel file on Nextcloud"),
            ("player export", "Export active PLTE players to Ladys.xlsx"),
            ("player import", "Import active PLTE players from Ladys.xlsx"),
            ("donations export", "Export donations sheet for active PLTE players"),
            ("donations import", "Import donations from Donations.xlsx"),
        ],
    )


def handle_command(command, args):
    if is_help_request(command, *args):
        print_help()
        return

    handlers = {
        "create": _handle_create,
        "import": _handle_import,
        "player": _handle_player,
        "donations": _handle_donations,
    }
    handler = handlers.get(command)
    if handler is None:
        print_unknown_command("sheet", command)
        print_help()
        return
    handler(args)


def _parse_match_id_arg(args, usage: str):
    if len(args) != 1:
        print(usage)
        return None
    match_id = parse_int(args[0], default=None)
    if match_id is None:
        print("❌ Match ID must be an integer.")
        return None
    return match_id


def _handle_create(args):
    match_id = _parse_match_id_arg(args, USAGE_SHEET_CREATE)
    if match_id is None:
        return

    with connect_db() as conn:
        match = get_match_info(conn, match_id)
        if not match:
            print("❌ No match found.")
            return

        _, _, season, _, _ = match
        ranked_players = rank_active_plte_for_season(conn, season) or get_active_players(conn)

    url, uploaded = generate_excel(match, ranked_players, output_path=NEXTCLOUD_BASE)
    print(f"✅ {url} ({'Created' if uploaded else 'Already existed'})")


def _handle_import(args):
    match_id = _parse_match_id_arg(args, USAGE_SHEET_IMPORT)
    if match_id is None:
        return
    import_excel_to_matchscore(match_id)


def _handle_player(args):
    if not args:
        print(USAGE_SHEET_PLAYER)
        return
    sub = args[0]
    if sub == "export":
        export_players_to_excel()
    elif sub == "import":
        import_players_from_excel()
    else:
        print(USAGE_SHEET_PLAYER)


def _handle_donations(args):
    if not args:
        print(USAGE_SHEET_DONATIONS)
        return
    sub = args[0]
    if sub == "export":
        export_donations_to_excel()
    elif sub == "import":
        import_donations_from_excel()
    else:
        print(USAGE_SHEET_DONATIONS)


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2 or sys.argv[1] != "sheet":
        print_help()
    else:
        handle_command(sys.argv[2] if len(sys.argv) > 2 else "", sys.argv[3:])
