from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from hcr2.services.sheets import parse_k_amount, to_k


def save_workbook(workbook: Workbook, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def delete_local_file(path: Path) -> bool:
    try:
        path.unlink()
        return True
    except OSError:
        return False


def autofit_columns(ws, min_w: int = 10, max_w: int = 60) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col_idx, col in enumerate(
        ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column),
        start=1,
    ):
        max_len = 0
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(min_w, min(max_w, max_len + 2))


def build_players_workbook(export_columns: list[str], rows: list[tuple]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "players"

    ws.append(export_columns)
    for row in rows:
        ws.append(list(row))

    autofit_columns(ws, min_w=10, max_w=60)
    return wb


def build_donations_workbook(rows: list[tuple[int, str, int]], today: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "donations"

    ws["A1"] = "Date:"
    ws["A2"] = today

    ws["A3"] = "id"
    ws["B3"] = "name"
    ws["C3"] = "donation (k)"
    ws["D3"] = "previous (k)"

    for cell in ("A3", "B3", "C3", "D3"):
        ws[cell].font = Font(bold=True)

    row_idx = 4
    for player_id, name, previous in rows:
        ws.cell(row=row_idx, column=1, value=player_id)
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value="")
        ws.cell(row=row_idx, column=4, value=to_k(previous))
        row_idx += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    return wb


def build_match_sheet_workbook(
    match: tuple[int, str, int, str, str],
    players: list[tuple[int, str, str | None, str | None]],
    *,
    is_absent_on,
) -> Workbook:
    match_id, match_date_str, season, opponent, event = match

    wb = Workbook()
    ws = wb.active
    ws.title = "Match Info"

    ws.append([f"Match ID: {match_id}", f"Date: {match_date_str}", f"Season: {season}", f"Opponent: {opponent}", f"Event: {event}"])

    ws.insert_rows(2, amount=1)
    ws["A2"] = "Result"
    ws["B2"] = "Power Ladies -->"
    ws["C2"] = ""
    ws["D2"] = ""
    ws["E2"] = f"<-- {opponent}"

    ws.append(["MatchID", "PlayerID", "Player", "Score", "Points", "Absent", "Checkin", "Notes"])

    for row in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=7):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 9
    ws.column_dimensions["H"].width = 130

    ws["H3"] = (
        "H1: Did not drive: enter Score=0 and Points=0.\n"
        "H2: Set Absent to true when a player is excused (vacation etc.).\n"
        "H3: Set Checkin to true when a player logged into the match but did not drive.\n"
        "H4: If a player left the team but is still listed, delete the row.\n"
        "H5: If a player is missing, add them with the correct ID.\n"
        "H6: If a missing player has not been created yet, enter 'a' for add in column B instead of the ID. The player is created during import.\n"
        "H7: Enter the match results in cell C2 (Ladies) and D2 (opponent).\n"
        "H8: Column C (Player) may be corrected when someone changed their name; the new name is stored during import. Leave it as it is otherwise, and never use it for notes — use column H."
    )
    ws["H3"].alignment = Alignment(wrap_text=True, vertical="top")

    for pid, name, away_from, away_until in players:
        absent_flag = is_absent_on(match_date_str, away_from, away_until)
        ws.append([match_id, pid, name, "", "", "true" if absent_flag else "false", "", ""])

    align_center = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = align_center

    return wb


def read_players_workbook(path: Path) -> tuple[list[str], list[dict[str, Any]]] | tuple[None, None]:
    wb = load_workbook(filename=path, data_only=True)
    ws = wb.active

    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    header = [str(c) if c is not None else "" for c in (first_row or [])]
    header = [h.strip() for h in header]
    if not header or "id" not in header:
        return None, None

    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        rows.append(dict(zip(header, row)))
    return header, rows


def read_donations_workbook(path: Path) -> tuple[str | None, list[tuple[int, int]], int]:
    wb = load_workbook(filename=path, data_only=True)
    ws = wb.active

    date_str = _read_donation_date(ws["A2"].value)
    if date_str is None:
        return None, [], 0

    entries: list[tuple[int, int]] = []
    errors = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row:
            continue
        pid_val = row[0]
        donation_val = row[2] if len(row) >= 3 else None

        if pid_val is None:
            continue
        pid_int = _read_int(pid_val)
        if pid_int is None:
            errors += 1
            continue

        if donation_val is None or (isinstance(donation_val, str) and donation_val.strip() == ""):
            continue

        donation_int = parse_k_amount(donation_val)
        if donation_int is None:
            errors += 1
            continue
        entries.append((pid_int, donation_int))

    return date_str, entries, errors


def read_match_sheet_workbook(path: Path) -> tuple[int | None, int | None, list[tuple[int, tuple]]]:
    wb = load_workbook(filename=path, data_only=True)
    ws = wb.active
    lady_score = _read_int(ws["C2"].value)
    opponent_score = _read_int(ws["D2"].value)
    rows = [
        (row_idx, row)
        for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4)
        if row and len(row) >= 3
    ]
    return lady_score, opponent_score, rows


def _read_donation_date(value) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        return value.strip()
    return None


def _read_int(value) -> int | None:
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    return None
